import openpyxl
import re
import tkinter as tk
from tkinter import filedialog
import os

# Функция для поиска адресов электронной почты в строке
def find_emails_in_string(text):
    email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    return re.findall(email_pattern, text)

# Создаем графический интерфейс для выбора файла
root = tk.Tk()
root.withdraw()

file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])

# Открываем выбранный xlsx файл
wb = openpyxl.load_workbook(file_path)

emails = []

# Перебираем все листы в файле
for sheet in wb.sheetnames:
    current_sheet = wb[sheet]
    
    # Проходим по всем ячейкам и ищем адреса электронной почты
    for row in current_sheet.iter_rows(values_only=True):
        for cell in row:
            if cell and isinstance(cell, str):
                emails += find_emails_in_string(cell)

# Создаем txt файл для сохранения адресов
output_file = os.path.splitext(file_path)[0] + "_emails.txt"

with open(output_file, 'w') as file:
    for email in emails:
        file.write(email + '\n')

print(f"Адреса электронной почты сохранены в файле: {output_file}")

# Закрываем файл
wb.close()
